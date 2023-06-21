from qgis.core import *
from PyQt5.QtWidgets import *
import qgis
from qgis.PyQt.QtCore import *
import processing
import openpyxl
import os
from processing.core.Processing import Processing
Processing.initialize()

def le_todos_os_layers():
        '''
        lê todos os layers do painel do QGIS
        e retorna os ids e nomes
        '''
        registry = QgsProject.instance()
        layers_names = []
        layers=[]
        for layer in QgsProject.instance().mapLayers().values():
            layers_names.append(layer.name())
        for i in range(len(layers_names)):
            name=layers_names[i]
            layers.append(registry.mapLayersByName( name )[0])
        return(layers_names,layers)

def retornar_layer(nome):
        '''
        retorna o layer pelo nome
        '''
        registry = QgsProject.instance()
        return registry.mapLayersByName( nome )[0]
def le_atributos(layer):
    # get the feature's attributes
    attrs = layer.attributes()
    return attrs
def buffer(layer,raio):
    result=processing.run("qgis:buffer", {
            'INPUT':layer, 
            'DISTANCE':raio,
            'SEGMENTS:':5,
            'END_CAP_STYLE':0,
            'JOIN_STYLE':0,
            'MITER_LIMIT':2,
            'DISSOLVE':False,
            'OUTPUT':'memory:'})["OUTPUT"]
    return result
def identifica_sobreposicoes(layer1,layer2,area_turbina):
    matriz_sobreposicao=[]
    for i in layer1.getFeatures():
        matriz_sobreposicao.append([])
        for f in layer2.getFeatures():
                # the geometry type can be of single or multi type
                #print(le_atributos(f))
                if i.geometry().intersects(f.geometry()):
                    index = area_turbina[0].index(le_atributos(i)[1])
                    if area_turbina[1][index] != le_atributos(f)[1] :
                        matriz_sobreposicao[count].append(le_atributos(f)[1])
    return matriz_sobreposicao
def retorna_sobreposicao(layer1,layer2,area_turbina):
    #os layers de pontos precisam ser reprojetados ou os layers de parques 
    matriz_sobreposicao=[]
    count=0
    for i in layer1.getFeatures():
        matriz_sobreposicao.append([])
        for f in layer2.getFeatures():
                # the geometry type can be of single or multi type
                if i.geometry().intersects(f.geometry()):
                    index = area_turbina[0].index(le_atributos(i)[1])
                    if area_turbina[1][index] != le_atributos(f)[1] :
                        matriz_sobreposicao[count].append(le_atributos(f)[1])
        count=count+1
    return matriz_sobreposicao
def join_by_location(layer1,layer2):
    result=processing.run("qgis:joinattributesbylocation", {
            'INPUT':layer1, 
            'JOIN':layer2,
            'PREDICATE':[0],
            'JOIN_FIELDS':[],
            'METHOD':0,
            'DISCARD_NONMATCHING':False,
            'PREFIX':'',
            'OUTPUT':'memory:'})["OUTPUT"]
    return result
def list_duplicates(seq):
  seen = set()
  seen_add = seen.add
  # adds all elements it doesn't know yet to seen and all other to seen_twice
  seen_twice = set( x for x in seq if x in seen or seen_add(x) )
  # turn the set into a list (as requested)
  return list( seen_twice )

def ordena_lista(x,y):
    # ordene os índices em vez dos elementos em si
    indices = list(range(len(x)))
    indices.sort(key=lambda i: x[i]) # ordene os índices com relação ao seu respectivo valor em x

    # crie as listas baseado na ordem dos índices
    new_x = [x[i] for i in indices]
    new_y = [y[i] for i in indices]
    return new_x,new_y
def indices_valores_iguais(mylist,value):
    '''
    retorna indice de valores iguais ao solicitado na lista 
    '''
    return [i for i,x in enumerate(mylist) if x==value]
def diferenca(lyr1,lyr2):
    result=processing.run("qgis:difference", {
            'INPUT':lyr1, 
            'OVERLAY':lyr2,
            'OUTPUT':'memory:'})["OUTPUT"]
    return result
def salva_excel_existente(paths,aero_area,sobrep,list,nome_base,nome_aba,header_init):
            '''
            Salva numa planilha existente , de acordo com a linha e
            coluna , sem apagar o que havia antes da planilha
            '''
            name = paths
            #df = pandas.DataFrame(matriz_excel,columns=colunas)
            if os.path.isfile(name) :
                wb=openpyxl.load_workbook(filename=name)
            else:
                wb = openpyxl.Workbook()
                ws =  wb.active
                ws.title = "Aeros"
                colunas=['Projeto','Turbina','Área','Sobreposição 1','Sobreposição 2','Sobreposição 3','Sobreposição 4','Sobreposição 5','Sobreposição 6','Sobreposição 7','Sobreposição 8','Sobreposição 9','Sobreposição 10']
                for i in range(1,15):
                    ws.cell(row=1,column=i).value = colunas[i-1]
            #para xlsm wb=openpyxl.load_workbook(filename=name,read_only=False, keep_vba=True)
            ws = wb['Aeros']
            for i in range(2,len(list)+2):
                    ws.cell(row=i,column=1).value = nome_base
            for i in range(2,len(list)+2):
                    ws.cell(row=i,column=2).value = list[i-2]
            for i in range(2,len(aero_area[0])+2):
                ws.cell(row = i, column = 3).value= aero_area[0][i-2]
                ws.cell(row = i, column = 4).value= aero_area[1][i-2]
                if sobrep[i-2]!=[] :
                    for j in range(5,5+len(sobrep[i-2])):
                        ws.cell(row = i, column = j).value = sobrep[i-2][j-5]
            wb.save(name)
            
def main(raio,layer_prop,layer_aero,path):
    layer1=layer_aero
    layer2=layer_prop
    layer3=buffer(layer1,raio)
    resultado_aeros_area=join_by_location(layer1,layer2)
    resultado_buffers_area=join_by_location(layer3,layer2)
    #0 inteseccao
    area_turbina=[[],[]]
    fase=[]
    for i in resultado_aeros_area.getFeatures():
        area_turbina[0].append(le_atributos(i)[1])
        area_turbina[1].append(le_atributos(i)[4])
        fase.append(le_atributos(i)[2])
    new_area_turbina=[[],[]]
    (new_area_turbina[0],new_area_turbina[1])=ordena_lista(area_turbina[0],area_turbina[1])
    temp=[[],[]]
    for i in resultado_buffers_area.getFeatures():
        temp[0].append(le_atributos(i)[1])
        temp[1].append(le_atributos(i)[4])
    new_temp=[[],[]]
    (new_temp[0],new_temp[1])=ordena_lista(temp[0],temp[1])
    sobreposicoes=[[]]
    duplicatas=list_duplicates(new_temp[0])

    
    

    for i in range (len(new_area_turbina[0])):
        if new_area_turbina[0][i] in duplicatas:
            indexs=indices_valores_iguais( new_temp[0], new_area_turbina[0][i])
            for j in indexs:
                if new_temp[1][j] != new_area_turbina[1][i]: sobreposicoes[i].append(new_temp[1][j])
        sobreposicoes.append([])
    #verifica os buffers que estao para fora dos parques 
    buffers_ANM=diferenca(layer3,layer2)
    buffers_anm=[]
    for i in buffers_ANM.getFeatures():
        buffers_anm.append(le_atributos(i)[1])
    
    [ sobreposicoes[i].append('Anm') for i in range(len(new_area_turbina[0])) if new_area_turbina[0][i] in buffers_anm ]
    print(sobreposicoes)
    
    salva_excel_existente(path,new_area_turbina,sobreposicoes,fase,layer_prop,'Aeros',0)
    msg = QMessageBox()
    msg.setText("Planilha de saida criada.")
    msg.exec()
